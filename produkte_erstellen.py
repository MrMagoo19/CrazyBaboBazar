import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Produkte"

# Header
headers = [
    "Nr.", "Produktname (DE)", "Kurzbeschreibung", "Kategorie",
    "Geschätzter Preis (€)", "Amazon DE Suchbegriff", "Zielgruppe",
    "Highlight / USP", "Bewertung Viral-Potenzial (1-5)"
]

header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
alt_fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 100 Produkte im thisiswhyimbroke-Stil
products = [
    # --- KÜCHE & HAUSHALT ---
    (1, "Weinflaschen-Kopf-Halter", "Skull-förmiger Weinflaschenhalter aus Metall – hält die Flasche schräg wie ein Weinkenner", "Küche & Bar", "25–35", "Weinflaschenhalter Totenkopf Metall", "Weinliebhaber, Party-Hosts", "Perfektes Gesprächsstarter-Stück", 4),
    (2, "Taco-Halter-Board", "Holzbrett mit 6 Taco-Ständern für das perfekte Taco-Tuesday-Erlebnis", "Küche & Haushalt", "20–30", "Taco Halter Holz Ständer", "Foodie, Familien", "Instagram-tauglich, praktisch", 5),
    (3, "Ramen-Nudelschüssel mit Stäbchenhalter", "Große Keramikschüssel mit eingebautem Stäbchenauflage und Deckel", "Küche", "18–28", "Ramen Schüssel Keramik Stäbchen", "Anime-Fans, Kochbegeisterte", "Japanisches Design, spülmaschinenfest", 4),
    (4, "Eiswürfelform Tetris-Steine", "Silikonform für Tetris-förmige Eiswürfel – perfekt für Cocktails", "Küche & Bar", "12–18", "Eiswürfelform Tetris Silikon", "Gamer, Cocktail-Fans", "Nostalgie-Faktor, viral auf TikTok", 5),
    (5, "Pizza-Schere mit Spatel", "Edelstahlschere zum direkten Schneiden von Pizza auf dem Blech mit integriertem Heber", "Küche", "15–25", "Pizzaschere mit Spatel Edelstahl", "Pizzaliebhaber", "Kein Beschädigen der Backform", 4),
    (6, "Avocado-Aufbewahrungsbox", "Silikonbehälter in Avocadoform – hält aufgeschnittene Avocado frisch", "Küche", "8–15", "Avocado Aufbewahrung Silikonbox", "Gesundheitsbewusste, Meal-Prepper", "Witzig & funktional", 4),
    (7, "Dinosaurier-Sieb / Nudelsieb", "Großes Küchensieb in Form eines Dinosauriers – Wasser läuft durch den Mund", "Küche", "20–30", "Dinosaurier Sieb Nudelsieb Küche", "Familien, Kinder, Foodie-Eltern", "Viral-Design, Spaß beim Kochen", 5),
    (8, "Bacon-Wecker", "Wecker der nach Speck riecht und Bratgeräusche macht um sanft zu wecken", "Küche & Schlafzimmer", "40–60", "Bacon Wecker Geruch Speck", "Frühstücksfans, Gadget-Liebhaber", "Einzigartiges Weckerlebnis", 5),
    (9, "Kaffeetassen-Wärmer USB", "Elektrische Untersetzer-Wärmeplatte für Tassen via USB", "Büro & Küche", "18–28", "Tassenwärmer USB Untersetzer Heizplatte", "Büroarbeiter, Kaffeetrinker", "Kein kalter Kaffee mehr", 4),
    (10, "Spaghetti-Portionierer Holz", "Holzstab mit Löchern für 1–4 Portionen Spaghetti – nie mehr zu viel kochen", "Küche", "8–12", "Spaghetti Portionierer Holz Nudeln", "WG-Bewohner, Studenten", "Einfach & günstig, löst echtes Problem", 3),
    # --- TECH & GADGETS ---
    (11, "Fingerabdruck-Vorhängeschloss", "Schloss öffnet in 0,5 Sek per Fingerabdruck – kein Schlüssel nötig", "Tech & Sicherheit", "35–55", "Fingerabdruck Vorhängeschloss biometrisch", "Fahrradfahrer, Sportler", "High-Tech zu günstigem Preis", 4),
    (12, "Mini-Beamer Smartphone", "Taschendrucker-großer Pico-Projektor mit HDMI & USB-C für unterwegs", "Tech & Entertainment", "80–130", "Mini Beamer Pico Projektor USB-C", "Studenten, Reisende, Netflix-Fans", "Kino überall", 5),
    (13, "LED-Handschuhe mit Fingerlichtern", "Jeder Finger hat eine eigene LED – ideal für Partys und Konzerte", "Party & Fun", "10–18", "LED Handschuhe Finger Lichter Party", "Party-Gänger, Festivals", "TikTok-viral, günstig", 5),
    (14, "Intelligentes Pflanzenbewässerungs-Set", "Selbstbewässernde Kegelspikes aus Keramik – einfach in die Flasche stecken", "Garten & Zuhause", "12–20", "Selbstbewässerung Pflanzen Keramik Spike", "Pflanzenliebhaber, Reisende", "Rettet Pflanzen im Urlaub", 4),
    (15, "RGB Gaming Mauspad XXL", "Riesiges Schreibtisch-Mauspad mit RGB-Beleuchtung und USB-Hub", "Gaming & Tech", "25–45", "Gaming Mauspad XXL RGB LED USB", "Gamer, Streamer", "Desk-Setup Upgrade", 4),
    (16, "Kabellose Mini-Tastatur mit Touchpad", "Kompakte Bluetooth-Tastatur mit eingebautem Trackpad – ideal für Smart-TV", "Tech", "25–40", "Mini Tastatur Bluetooth Touchpad Smart TV", "Smart-TV-Nutzer, Couch-Surfer", "Endlich bequemes Tippen vom Sofa", 4),
    (17, "Handgelenk-Handgelenkband Fitness-Tracker Dummy", "Fitness-Tracker-Look mit tatsächlicher Funktion für unter 30€", "Tech & Fitness", "20–35", "Fitness Tracker Smartband Herzfrequenz günstig", "Fitness-Einsteiger", "Gutes Preis-Leistungs-Verhältnis", 3),
    (18, "Solar-Power-Bank", "Powerbank mit Solarpanel – lädt per Sonnenlicht nach", "Tech & Outdoor", "30–50", "Solar Powerbank Outdoor USB-C", "Camper, Festival-Besucher", "Endlos Energie in der Natur", 4),
    (19, "Magnetische Kabelhalter", "Selbstklebende Magnete halten Kabel ordentlich am Schreibtisch", "Tech & Organisation", "8–15", "Magnetischer Kabelhalter Schreibtisch Organizer", "Büroarbeiter, Ordnungsliebhaber", "Sofort wirksam, günstig", 3),
    (20, "USB-C Hub 7-in-1", "Kompakter Multiport-Adapter mit HDMI, SD-Karte, USB-A, USB-C", "Tech", "25–40", "USB-C Hub 7in1 HDMI MacBook", "MacBook-Nutzer, Studenten", "Unverzichtbar für modernes Arbeiten", 4),
    # --- ZIMMER & DEKO ---
    (21, "Neon-Schild LED personalisierbar", "LED-Neonschild mit individuellem Text – USB-betrieben", "Deko & Wohnen", "30–60", "LED Neon Schild personalisiert Wanddeko", "Interior-Fans, Teenager, Streamer", "Viraler Deko-Trend", 5),
    (22, "Mondlampe 3D", "Realistische 3D-gedruckte Mondkugel als Nachtlicht mit Farbwechsel", "Deko & Beleuchtung", "25–45", "3D Mondlampe Nachtlicht LED", "Raumfahrt-Fans, Romantiker", "Wow-Effekt garantiert", 5),
    (23, "Bücherregal Unsichtbar (schwebend)", "Wandhalterung die aussieht als würden Bücher in der Luft schweben", "Wohnen & Deko", "15–25", "Unsichtbares Bücherregal Wandmontage schwebend", "Buchliebhaber, Interior-Designer", "Magischer Effekt, einfache Montage", 5),
    (24, "Galaxy-Projektor Sternenhimmel", "Projiziert Galaxien, Sterne und Aurora-Effekte an Decke und Wände", "Deko & Schlafzimmer", "30–55", "Galaxie Projektor Sternenhimmel Nachtlicht", "Kinder, Romantiker, Entspannungs-Fans", "Traumhaftes Ambiente", 5),
    (25, "Aquarium-Kaktus LED", "Kleiner LED-Kaktus aus Glas mit buntem Licht – pflegeleichte Deko", "Deko", "15–25", "LED Kaktus Lampe Nachtlicht Deko", "Teenager, Mini-Apartment-Bewohner", "Null Pflege, maximaler Vibe", 4),
    (26, "Vinyl-Schallplatten-Wanduhr", "Uhr aus echter recycelter Schallplatte – Retro-Stil", "Deko & Wohnen", "20–35", "Schallplatten Wanduhr Vinyl Retro", "Musikfans, Vintage-Liebhaber", "Nachhaltiges Upcycling-Design", 4),
    (27, "Ozean-Wellen-Projektor", "Projiziert beruhigende Wellenmuster in Blau und Grün an die Wand", "Deko & Wellness", "25–40", "Ozean Wellen Projektor Nachtlicht Entspannung", "Entspannungs-Fans, Meeres-Liebhaber", "Spa-Atmosphäre zu Hause", 4),
    (28, "Leuchtender Pflanzentopf Solar", "Blumentopf mit eingebautem Solarlicht leuchtet nachts", "Garten & Deko", "20–35", "Solar Blumentopf LED Garten Nachtlicht", "Gartenbesitzer, Balkon-Fans", "Automatisch, kein Strom nötig", 3),
    (29, "Hologramm-Fan Display", "3D-Hologramm-Lüfter zeigt animierte 3D-Objekte in der Luft", "Tech & Deko", "60–100", "Hologramm Fan Display 3D LED", "Gadget-Enthusiasten, Shop-Besitzer", "Echter Hingucker und Gesprächsstarter", 5),
    (30, "Wolken-Lampe mit Blitz-Effekt", "Aufblasbare Wolke mit LED-Blitz simuliert Gewitter – Ambilight", "Deko & Beleuchtung", "35–60", "Wolken Lampe Gewitter Blitz LED Deko", "Streamer, Interior-Fans", "Extremes Eye-Catcher, TikTok viral", 5),
    # --- OUTDOOR & ABENTEUER ---
    (31, "Hängematte für 2 Personen ultraleicht", "Ultraleichte Nylon-Hängematte für 2 Personen, packt sich mini klein", "Outdoor & Camping", "30–50", "Hängematte 2 Personen ultraleicht Camping Nylon", "Camper, Wanderer, Festival-Fans", "Passt in jede Rucksacktasche", 4),
    (32, "Wasserdichter Beutel zum Umhängen", "IPX8-zertifizierter Drybag mit Touch-Screen-Fenster fürs Handy", "Outdoor & Wassersport", "15–25", "Wasserdichter Beutel Handy Drybag IPX8", "Schwimmer, Kajak-Fahrer, Festival-Besucher", "Handy am Strand oder Bootsfahrt sichern", 4),
    (33, "Kompakter Klappstuhl ultraleicht", "Faltstuhl aus Aluminium, nur 500g schwer und Platz für 120kg", "Outdoor & Camping", "25–40", "Klappstuhl ultraleicht Aluminium Camping", "Wanderer, Outdoor-Fans, Festivals", "Passt in jeden Rucksack", 3),
    (34, "Stirnlampe mit Rotlichtmodus", "Leistungsstarke Stirnlampe mit Rotlicht für Nachtsichterhalt", "Outdoor & Camping", "20–35", "Stirnlampe Rotlicht Nachtsicht USB-C aufladbar", "Camper, Jogger, Kletterer", "Militärischer Rotlichtmodus", 3),
    (35, "Multitool Taschenmesser 16-in-1", "Edelstahl-Multitool mit Messer, Säge, Zange und mehr", "Outdoor & EDC", "25–45", "Multitool Taschenmesser Edelstahl 16in1", "Outdoor-Fans, Handwerker, Reisende", "Klassiker, immer nützlich", 4),
    (36, "Trinkflasche mit Wasserfilter integriert", "Filtert Bakterien & Parasiten direkt beim Trinken aus jedem Gewässer", "Outdoor & Survival", "35–55", "Trinkflasche Wasserfilter Camping Survival", "Wanderer, Survival-Fans, Reisende", "Trinkwasser überall auf der Welt", 4),
    (37, "Notfall-Paracord-Armband", "Armband aus 3m Paracord mit Pfeife, Kompass und Feuerstarter", "Outdoor & Survival", "10–18", "Paracord Armband Survival Kompass Pfeife", "Outdoor-Fans, Scouts, Abenteurer", "10 Funktionen am Handgelenk", 4),
    (38, "Camping-Hängematte mit Moskitonetz", "Ultraleichte Hängematte mit integriertem Insektenschutz-Netz", "Outdoor & Camping", "40–65", "Hängematte Moskitonetz Camping ultraleicht", "Camper, Dschungel-Reisende", "Schläft sicher in der Wildnis", 4),
    (39, "LED-Campinglaterne faltbar solar", "Faltbare Solarlaterne, aufladbar per USB, 360° Licht", "Outdoor & Camping", "15–25", "Solarlaterne faltbar LED Camping USB", "Camper, Balkon-Fans, Notfallausrüstung", "Kompakt, langlebig, solar", 4),
    (40, "Feuerstarter Ferrocerium Flintstein", "Magnesium-Feuerstein mit 12.000 Schlägen Lebensdauer", "Outdoor & Survival", "8–15", "Ferrocerium Feuerstein Feuerstarter Survival", "Camper, Survival-Fans", "Funktioniert auch bei Regen", 3),
    # --- GAMING & GEEK ---
    (41, "Retro Handheld Spielekonsole (400+ Spiele)", "Kompakte Handheld-Konsole mit vorinstallierten Retro-Spielen", "Gaming", "25–45", "Retro Handheld Spielekonsole Mini vorinstalliert", "Gamer, 80s/90s-Nostalgiker", "Hunderte Spiele, kein Internet nötig", 5),
    (42, "Arcade-Joystick für PC/PS4/Switch", "Klassischer Arcade-Stick für Kampfspiele, kompatibel mit mehreren Plattformen", "Gaming", "45–80", "Arcade Joystick Fight Stick PC PS4 Switch", "Fighting-Game-Fans", "Authentisches Arcade-Gefühl zuhause", 4),
    (43, "LED-Würfel Smart Deko", "Farbwechselnder LED-Würfel reagiert auf Musik oder Timer", "Gaming & Deko", "35–55", "LED Würfel Smart Deko Farbwechsel Musik", "Gamer, Deko-Fans", "Doppelt nützlich: Deko + Gadget", 4),
    (44, "Gamepad-Halter Schreibtisch", "Aluminium-Ständer für Controller, hält PS5/Xbox/Nintendo sauber", "Gaming & Organisation", "15–25", "Controller Halter Ständer Schreibtisch Aluminium PS5", "Gamer", "Ordnung im Gaming-Setup", 3),
    (45, "Tastatur-Handballenauflage Memory Foam", "Ergonomische Handballenauflage aus Memory-Foam für Gaming-Tastaturen", "Gaming & Ergonomie", "15–25", "Handballenauflage Tastatur Memory Foam Gaming", "Gamer, Büroarbeiter", "Reduziert RSI und Handgelenksschmerzen", 3),
    (46, "Pixel-Art LED-Panel programmierbar", "64x64 LED-Matrix zum Selbst-Programmieren von Pixel-Art und Animationen", "Gaming & DIY", "40–70", "LED Matrix Pixel Art Panel programmierbar", "Gamer, Maker, Kreative", "DIY-Spaß mit Wifi-Steuerung", 5),
    (47, "VR-Brille für Smartphone", "Cardboard-Stil VR-Headset für alle Smartphones bis 6,7 Zoll", "Gaming & VR", "20–35", "VR Brille Smartphone Headset Cardboard", "Technik-Neugierige, Gamer", "VR-Einstieg für kleines Geld", 4),
    (48, "Gamer-Stuhl Kissen Lordose + Nacken", "Orthopädisches Kissen-Set für Gaming-Stühle und Bürostühle", "Gaming & Ergonomie", "20–35", "Gamerstuhl Kissen Set Lordose Nackenkissen", "Gamer, Büroarbeiter", "Rücken retten ohne neuen Stuhl", 3),
    (49, "Schreibtisch-Organizer Gaming RGB", "Bambus-Schreibtisch-Organizer mit RGB-Streifen unten dran", "Gaming & Organisation", "25–40", "Schreibtisch Organizer Bambus RGB Gaming", "Gamer, Streamer", "Ästhetisches Setup für wenig Geld", 4),
    (50, "Retro-Bartender Cocktail-Maschine Mini", "Miniatur-Bar-Setup mit Messbecher, Rezeptheft und Bar-Werkzeug", "Bar & Gaming", "30–50", "Cocktail Set Bar Werkzeug Starter Kit Barkeeper", "Cocktail-Fans, Gamer, Party-Hosts", "Mixologist-Starter-Kit", 4),
    # --- FASHION & ACCESSOIRES ---
    (51, "Unsichtbare Socken Anti-Rutsch", "Sehr flache Liner-Socken die im Sneaker unsichtbar bleiben", "Mode & Accessoires", "10–18", "Unsichtbare Socken Sneaker Liner Anti-Rutsch", "Sneaker-Fans", "Löst das leidige Socken-sieht-man Problem", 3),
    (52, "Regenbogen-Reflektions-Jacke", "Jacke mit Holographic-Stoff der bei Licht alle Farben reflektiert", "Mode & Outdoor", "50–90", "Holographic Jacke Regenbogen Reflektion Festival", "Festival-Fans, Fashion-Lovers", "Social-Media-Magnet", 5),
    (53, "Magnet-Portemonnaie ultraslim", "Ultradünnes Kartenetui mit RFID-Schutz und Pop-Up-Mechanismus", "Accessoires & Tech", "20–35", "Slim Wallet RFID Schutz Kartenhalter Magnet", "Minimalisten, Stadtmenschen", "Ersatz für dicke Geldbörse", 4),
    (54, "LED-Lauf-Schuhe leuchtend", "Sneaker mit LED-Sohlen die beim Laufen aufleuchten", "Mode & Fun", "35–65", "LED Schuhe leuchtend Sneaker Kinder Erwachsene", "Kinder, Party-Menschen, Extravagante", "Hingucker bei Nacht", 5),
    (55, "Smartwatch-Ladestation Bambus", "Eleganter Bambus-Ständer lädt Apple Watch und AirPods gleichzeitig", "Tech & Accessoires", "20–35", "Ladestation Bambus Apple Watch AirPods Ständer", "Apple-Nutzer, Minimalisten", "Ordnung auf dem Nachttisch", 4),
    (56, "Fidget-Ring aus Edelstahl", "Drehbarer Ring aus gebürstetem Edelstahl – Anxiety-Ring", "Accessoires & Wellness", "12–22", "Fidget Ring Edelstahl Angst Ring Spinner", "Stressgeplagte, Studierende", "Diskrete Anti-Stress-Hilfe", 4),
    (57, "Rucksack mit USB-Ladeport", "Wasserabweisender Laptop-Rucksack mit integriertem USB-A Ladeport", "Mode & Tech", "35–60", "Rucksack USB Ladeport Laptop wasserdicht", "Pendler, Studenten, Reisende", "Handy immer aufgeladen unterwegs", 4),
    (58, "Handschuhe Touchscreen-tauglich", "Strickhandschuhe deren alle Finger Touchscreens bedienen können", "Mode & Tech", "12–20", "Touchscreen Handschuhe Winter Smartphone", "Smartphone-Nutzer, Pendler", "Kein Ausziehen der Handschuhe mehr", 3),
    (59, "Magnetische Brillenhalterung Auto", "Magnet-Clip fürs Auto-Armaturenbrett hält Brillen sicher", "Auto & Accessoires", "8–14", "Brillenhalter Auto Magnet Clip Armaturenbrett", "Brillenträger, Autofahrer", "Brillen immer griffbereit", 3),
    (60, "Mini-Parfüm-Zerstäuber nachfüllbar", "Metallischer Reise-Atomizer für unterwegs – 5ml fassend", "Accessoires & Beauty", "10–18", "Parfüm Zerstäuber Reise nachfüllbar Mini Atomizer", "Reisende, Business-Menschen", "Lieblingsparfüm immer dabei", 3),
    # --- WELLNESS & ENTSPANNUNG ---
    (61, "Massagepistole Faszienpistole Mini", "Kompakte Faszienpistole mit 4 Aufsätzen für Muskelentspannung", "Wellness & Fitness", "40–70", "Massagepistole Mini Faszien Muskel Entspannung", "Sportler, Büroarbeiter, Rückenschmerzgeplagte", "Professionelle Massage zu Hause", 5),
    (62, "Akupressurmatte Igel-Set", "Akupressurmatte + Kissen + Tasche für tägliche Entspannung", "Wellness", "25–45", "Akupressurmatte Set Kissen Entspannung Rücken", "Stressgeplagte, Yogafans", "Günstige Alternative zur Massage", 4),
    (63, "Schlafmaske mit Bluetooth Kopfhörern", "Weiche Schlafmaske mit eingebauten flachen BT-Kopfhörern zum Einschlafen", "Wellness & Schlaf", "25–40", "Schlafmaske Bluetooth Kopfhörer eingebaut Einschlafen", "Schlechte Schläfer, Reisende", "Musik + Dunkelheit = perfekter Schlaf", 5),
    (64, "Schaumstoff-Fußmassage-Roller", "Fußrolle mit Noppen für selbst-Reflexzonenmassage", "Wellness", "12–20", "Fußrolle Massage Reflexzonen Noppen Entspannung", "Stehberufe, Läufer", "Günstige Fußpflege täglich", 3),
    (65, "Aromatherapie Diffuser Ultraschall", "Ultraschall-Luftbefeuchter mit LED und ätherischen Ölen für 150qm", "Wellness & Deko", "25–45", "Aromatherapie Diffuser Ultraschall LED Luftbefeuchter", "Wellness-Fans, Zuhause-Worker", "Spa-Atmosphäre und bessere Luft", 4),
    (66, "Yoga-Rad Rücken Stretch", "Kunststoff-Yoga-Rad für tiefe Rückendehnungen und Herzöffner", "Wellness & Fitness", "30–50", "Yoga Rad Wheel Rücken Dehnung Stretch", "Yogis, Rückenschmerzgeplagte", "Intensivere Dehnungen möglich", 4),
    (67, "Kältetherapie Eisrolle für Gesicht", "Kühlende Gesichtsrolle aus rostfreiem Stahl – reduziert Schwellungen", "Beauty & Wellness", "12–20", "Eisrolle Gesicht Kältetherapie Gesichtsrolle Stahl", "Beauty-Fans, Morgen-Muffel", "Wach werden + Anti-Aging-Effekt", 4),
    (68, "Infrarot-Sauna-Decke", "Einwickelbare Infrarot-Sauna-Decke für Heimgebrauch", "Wellness", "80–140", "Infrarot Sauna Decke Heimsauna Entspannung", "Wellness-Fans, Detox-Begeisterte", "Sauna ohne Sauna-Zimmer", 5),
    (69, "Nackenmassagegerät Schulter Shiatsu", "Elektrisches Nacken-Schulter-Massagegerät mit Wärmefunktion", "Wellness", "35–60", "Nackenmassagegerät Shiatsu Schulter Wärme elektrisch", "Büroarbeiter, Stressgeplagte", "Professionelle Massage jeden Tag", 5),
    (70, "Meditationskissen Zafu Rund", "Rundes Buckwheat-gefülltes Meditationskissen für stabiles Sitzen", "Wellness & Yoga", "25–40", "Meditationskissen Zafu Buchweizen Meditation Yoga", "Meditierende, Yogafans", "Korrekte Sitzhaltung für tiefe Meditation", 3),
    # --- SPASS & PARTY ---
    (71, "Riesige Jenga-Set XXL", "Outdoor-XXL-Jenga aus Holzblöcken – Klötze bis 90cm hoch stapelbar", "Spiele & Party", "45–80", "Riesen Jenga XXL Outdoor Holz Gartenspiel", "Familien, Party-Hosts, Garten-Besitzer", "Bestes Gartenspiel für Gruppen", 5),
    (72, "Wein-Ping-Pong-Set", "Beer-Pong-Set mit Weingläsern und Weinpong-Regeln", "Party & Bar", "25–40", "Weinpong Set Weingläser Beer Pong Party", "Erwachsenen-Parties, Wein-Fans", "Erwachsener Twist auf Klassiker", 4),
    (73, "Kartenspiel 'What Do You Meme' DE", "Virales Meme-Kartenspiel auf Deutsch für 3-20 Spieler", "Spiele & Party", "20–30", "What Do You Meme Kartenspiel Deutsch Party", "Millennials, Gen Z, Partygruppen", "Internet-Kultur als Spiel", 5),
    (74, "Glow-in-the-Dark Frisbee", "Im Dunkeln leuchtende Frisbee-Scheibe mit LED-Ring", "Sport & Party", "12–20", "Leuchtende Frisbee Nacht LED Glow Dark", "Outdoor-Fans, Abend-Sportler", "Sport auch nach Sonnenuntergang", 4),
    (75, "Tisch-Shuffleboard Mini", "45cm Tisch-Shuffleboard aus Holz für Partys und Familien", "Spiele & Party", "30–50", "Tisch Shuffleboard Mini Holz Familienspiel Party", "Familien, Party-Menschen", "Wenig bekannt, macht süchtig", 4),
    (76, "Waterpong-Wasserballon-Set", "Wasserpistolen + Wasserballons + Pong-Becher = ultimatives Sommer-Set", "Party & Outdoor", "15–25", "Wasserballon Wasserpistole Set Sommer Pong", "Kinder, Familien, Sommer-Parties", "Spaß für alle Altersgruppen", 4),
    (77, "Kicker Tischfußball Mini Reise", "Klappbarer Mini-Kicker für Schreibtisch oder unterwegs", "Spiele & Fun", "20–35", "Mini Kicker Tischfußball Reise klappbar Büro", "Büro-Kollegen, Studenten", "Mittags-Pause nie wieder langweilig", 4),
    (78, "Dartscheibe mit LED-Licht", "Elektronische Dartscheibe mit Soundeffekten und Punktestand", "Spiele & Party", "35–65", "Dartscheibe elektronisch LED Sound Punkte", "Spiele-Fans, Man-Cave-Besitzer", "Professioneller Dart-Spaß zu Hause", 4),
    (79, "Würfelturm aus Holz Maßgeschneidert", "Handgefertigter Würfelturm verhindert Würfel-Verlust beim Brettspielen", "Spiele", "20–35", "Würfelturm Holz Brettspiel handgefertigt", "Brettspieler, Rollenspieler", "Must-have für Brettspiel-Abende", 3),
    (80, "Escape-Room Spiel 'Unlock!' DE", "Kartenbasiertes Escape-Room-Spiel für 1-6 Spieler auf Deutsch", "Spiele", "15–25", "Unlock Escape Room Kartenspiel Deutsch", "Rätsel-Fans, Freunde-Gruppen", "Escape-Room-Feeling ohne teure Buchung", 4),
    # --- AUTO & PENDLER ---
    (81, "Sitzheizung Auto universal", "Sitzauflage mit USB-Heizung für jeden Auto-Sitz", "Auto", "20–35", "Sitzheizung Auto USB universal Auflage Heizung", "Autofahrer, Winterpendler", "Warmer Sitz ohne Neuwagenpreis", 4),
    (82, "Handyhalter Auto Magnet Lüftung", "Starker Magnet-Handyhalter für alle Smartphones, Lüftungsmontage", "Auto & Tech", "10–18", "Handyhalter Auto Magnet Lüftung universal", "Autofahrer, GPS-Nutzer", "Kein wackeln, kein nerviges Einstellen", 4),
    (83, "Auto-Staubsauger Nass & Trocken", "Kompakter Auto-Staubsauger mit 12V-Anschluss und Nassfunction", "Auto", "20–35", "Auto Staubsauger Nass Trocken 12V kompakt", "Autobesitzer", "Innenraum blitzblank in 5 Minuten", 3),
    (84, "Lenkrad-Tabletttisch", "Klapptisch für Lenkrad-Halterung – ideal für Pausen auf Autobahnraststätten", "Auto & Pendler", "15–25", "Lenkrad Tabletttisch Klapptisch Auto Essen Laptop", "Langstreckenfahrer, Trucker", "Büro im Auto", 4),
    (85, "Kfz-Ladegerät USB-C 65W", "Super-schnelllader für Auto: 65W USB-C + 18W USB-A gleichzeitig", "Auto & Tech", "20–35", "Kfz Ladegerät 65W USB-C Schnellladen Auto", "Elektro-Gadget-Nutzer, Vielfahrer", "Schnellstes Aufladen im Auto", 4),
    # --- BÜRO & PRODUKTIVITÄT ---
    (86, "Stehtisch-Aufsatz Converter", "Höhenverstellbarer Schreibtisch-Aufsatz für Sitz-Steh-Wechsel", "Büro & Ergonomie", "80–150", "Schreibtisch Aufsatz höhenverstellbar Stehtisch Converter", "Home-Office-Arbeiter, Büroarbeiter", "Gesünder arbeiten ohne teuren Stehtisch", 4),
    (87, "Schreibtisch-Organizer Bambus", "Modularer Bambus-Organizer mit Schubladen und Stifthalter", "Büro & Organisation", "25–45", "Bambus Schreibtisch Organizer Schublade Stifthalter", "Ordnungsliebhaber, Home-Office", "Natürliches Material, schönes Design", 3),
    (88, "Ergonomische Vertikalmaus", "Vertikale Maus reduziert Drehung des Unterarms bei Benutzung", "Büro & Ergonomie", "25–45", "Vertikalmaus ergonomisch kabellos USB-C", "Büroarbeiter, Schreibtisch-Jobber", "Reduziert Mausarm/Sehnenscheidenentzündung", 4),
    (89, "Whiteboard-Folie selbstklebend", "Transparente selbstklebende Folie macht jede Wand zum Whiteboard", "Büro & Home-Office", "20–35", "Whiteboard Folie selbstklebend transparent Wand", "Kreative, Lehrer, Startup-Gründer", "Whiteboard überall ohne Bohren", 4),
    (90, "Pomodoro-Timer analoger Küchentimer", "Mechanischer Tomaten-Timer für 25-Min-Pomodoro-Technik", "Büro & Produktivität", "10–18", "Pomodoro Timer Küchentimer Tomate mechanisch", "Studenten, Prokrastinierer", "Psychologisch wirksame Zeitmethode", 4),
    # --- BIZARRE & WTF ---
    (91, "Gummiente XXL 1 Meter", "Riesige 1-Meter-Badeente aus Gummi – für Badewanne oder Pool", "Fun & Bizarre", "35–60", "Riesige Gummiente XXL 1 Meter Badewanne Pool", "WTF-Käufer, Spaßvögel, Pool-Besitzer", "Absurd groß, macht jeden glücklich", 5),
    (92, "Totenkopf-Eiswürfelform", "Silikonform macht Eiswürfel in Totenkopf-Form – für Halloween & Bar", "Küche & Fun", "10–18", "Totenkopf Eiswürfel Form Silikon Skull Bar", "Bar-Fans, Halloween-Liebhaber", "Gruselig-coole Party-Zutat", 5),
    (93, "Nixie-Röhren-Uhr Bausatz", "Retro-Nixie-Röhren-Uhr zum Selbst-Löten – für Tech-Nostalgiker", "Tech & DIY", "60–100", "Nixie Röhren Uhr Bausatz Retro DIY löten", "Bastler, Retro-Tech-Fans", "Einzigartiges Kunstobjekt das Technik zeigt", 5),
    (94, "Wackeldackel modern LED", "Moderner Wackeldackel mit LED-Augen und USB-Anschluss", "Fun & Deko", "15–25", "Wackeldackel LED modern Deko Auto Schreibtisch", "Nostalgie-Fans, Spaßvögel", "Klassiker modernisiert", 4),
    (95, "Kaugummi-Spender Gadget", "Schreibtisch-Kaugummi-Dispenser der nach Wurfeinwurf einen gibt", "Fun & Büro", "15–25", "Kaugummi Spender Automat Schreibtisch Gadget", "Bürohumor-Fans, Naschkatzen", "Gesprächsstarter im Büro", 4),
    (96, "Spaghetti-Uhr Pasta-Timer", "Sanduhr genau für 8 Minuten – exakt richtig für Al-Dente-Pasta", "Küche & Fun", "12–20", "Pasta Timer Sanduhr Spaghetti 8 Minuten kochen", "Pasta-Fans, Küchen-Gadget-Liebhaber", "Witzig und tatsächlich nützlich", 4),
    (97, "Bananentelefonhörer für Smartphone", "Echter Bananenhörer-Aufsatz für Smartphones für witzige Gespräche", "Fun & Tech", "12–20", "Bananenhörer Telefon Smartphone Aufsatz witzig", "Bürowiticze, Spaßvögel", "Foto-Gag und echter Hörer zugleich", 5),
    (98, "Pizza-Socken 3D-Optik", "Socken mit 3D-aufgedruckter echter Pizzascheibe – ultrarealistische Optik", "Mode & Fun", "8–15", "Pizza Socken 3D Optik lustig Geschenk", "Pizzafans, Beschenkende ohne Idee", "Perfektes Last-Minute-Geschenk", 5),
    (99, "Katzenpfoten-Handschuhe zum Streicheln", "Plüsch-Handschuhe in Katzenpfoten-Form – zum gegenseitigen Streicheln", "Fun & Pet", "12–20", "Katzenpfoten Handschuhe Plüsch Streicheln Pet", "Katzenbesitzer, Spaßvögel", "Internet-Hit, unwiderstehlich absurd", 5),
    (100, "Einhorn-Kopfbedeckung Poncho-Decke", "Einhorn-Kigurumi-Kapuzendecke für Erwachsene mit Einhorn-Horn", "Fun & Wohnen", "25–40", "Einhorn Decke Kapuze Kigurumi Poncho Erwachsene", "Einhorn-Fans, Couch-Potatoes", "Viral-Potenzial durch pure Niedlichkeit", 5),
]

cat_colors = {
    "Küche & Bar": "FFE0B2", "Küche & Haushalt": "FFE0B2", "Küche": "FFE0B2",
    "Küche & Schlafzimmer": "FFE0B2", "Küche & Fun": "FFE0B2",
    "Tech & Sicherheit": "B3E5FC", "Tech & Entertainment": "B3E5FC",
    "Tech": "B3E5FC", "Tech & Outdoor": "B3E5FC",
    "Tech & Organisation": "B3E5FC", "Gaming & Tech": "B3E5FC",
    "Party & Fun": "F8BBD9", "Spiele & Party": "F8BBD9",
    "Party & Bar": "F8BBD9", "Party & Outdoor": "F8BBD9", "Spiele": "F8BBD9",
    "Sport & Party": "F8BBD9",
    "Deko & Wohnen": "C8E6C9", "Deko & Beleuchtung": "C8E6C9",
    "Deko": "C8E6C9", "Deko & Schlafzimmer": "C8E6C9",
    "Deko & Wellness": "C8E6C9", "Wohnen & Deko": "C8E6C9",
    "Garten & Deko": "C8E6C9", "Tech & Deko": "C8E6C9",
    "Garten & Zuhause": "C8E6C9", "Garten & Zuhause": "C8E6C9",
    "Outdoor & Camping": "DCEDC8", "Outdoor & Wassersport": "DCEDC8",
    "Outdoor & Survival": "DCEDC8", "Outdoor & EDC": "DCEDC8",
    "Outdoor & Abenteuer": "DCEDC8",
    "Gaming": "EDE7F6", "Gaming & Organisation": "EDE7F6",
    "Gaming & Ergonomie": "EDE7F6", "Gaming & DIY": "EDE7F6",
    "Gaming & VR": "EDE7F6", "Gaming & Büro": "EDE7F6",
    "Wellness & Fitness": "FCE4EC", "Wellness": "FCE4EC",
    "Wellness & Schlaf": "FCE4EC", "Wellness & Yoga": "FCE4EC",
    "Beauty & Wellness": "FCE4EC",
    "Mode & Accessoires": "FFF9C4", "Mode & Outdoor": "FFF9C4",
    "Mode & Fun": "FFF9C4", "Mode & Tech": "FFF9C4",
    "Accessoires & Tech": "FFF9C4", "Accessoires & Wellness": "FFF9C4",
    "Accessoires & Beauty": "FFF9C4",
    "Auto": "CFD8DC", "Auto & Tech": "CFD8DC", "Auto & Pendler": "CFD8DC",
    "Büro & Ergonomie": "F3E5F5", "Büro & Organisation": "F3E5F5",
    "Büro & Home-Office": "F3E5F5", "Büro & Produktivität": "F3E5F5",
    "Fun & Bizarre": "FFCCBC", "Fun & Deko": "FFCCBC",
    "Fun & Büro": "FFCCBC", "Fun & Tech": "FFCCBC",
    "Fun & Pet": "FFCCBC", "Fun & Wohnen": "FFCCBC",
    "Tech & DIY": "FFCCBC", "Bar & Gaming": "FFCCBC",
}

thin = Side(style='thin', color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for row_data in products:
    row_num = row_data[0] + 1
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cat = row_data[3]
        fill_color = cat_colors.get(cat, "FFFFFF")
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if col == 9:  # Viral-Potenzial
            if val == 5:
                cell.font = Font(bold=True, color="C62828")
            elif val == 4:
                cell.font = Font(color="1B5E20")

# Spaltenbreiten
col_widths = [5, 30, 50, 22, 18, 38, 28, 35, 10]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Zeilenhöhe Header
ws.row_dimensions[1].height = 30

# Freeze header
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

out = "/home/batman/CrazyBaboBazar/produkte_thisiswhyimbroke_stil.xlsx"
wb.save(out)
print(f"Gespeichert: {out}")
