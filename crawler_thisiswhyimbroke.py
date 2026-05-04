"""
Crawler für thisiswhyimbroke.com
Ziel: 1000 Produkte mit allen verfügbaren Infos → Excel
"""
import asyncio
import json
import re
import time
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from crawl4ai import AsyncWebCrawler, CrawlerRunConfig, BrowserConfig
from crawl4ai.extraction_strategy import JsonCssExtractionStrategy
from bs4 import BeautifulSoup

OUTPUT_DIR = Path("/home/batman/CrazyBaboBazar")
BASE_URL = "https://www.thisiswhyimbroke.com"

# Alle bekannten Kategorieseiten + neue Seiten die wir entdecken
SEED_URLS = [
    f"{BASE_URL}/",
    f"{BASE_URL}/new/",
    f"{BASE_URL}/weird-stuff/",
    f"{BASE_URL}/cool-shit/",
    f"{BASE_URL}/gifts/",
    f"{BASE_URL}/cooking-gifts/",
    f"{BASE_URL}/cozy-gifts/",
    f"{BASE_URL}/organization-gifts/",
    f"{BASE_URL}/among-us-gifts/",
    f"{BASE_URL}/tech-gifts/",
    f"{BASE_URL}/funny-gifts/",
    f"{BASE_URL}/cheap-gifts/",
    f"{BASE_URL}/unique-gifts/",
    f"{BASE_URL}/cool-gadgets/",
    f"{BASE_URL}/gifts-for-men/",
    f"{BASE_URL}/gifts-for-women/",
    f"{BASE_URL}/gifts-for-him/",
    f"{BASE_URL}/gifts-for-her/",
    f"{BASE_URL}/gifts-for-dad/",
    f"{BASE_URL}/gifts-for-mom/",
    f"{BASE_URL}/gifts-for-teens/",
    f"{BASE_URL}/gifts-for-kids/",
    f"{BASE_URL}/stocking-stuffers/",
    f"{BASE_URL}/birthday-gifts/",
    f"{BASE_URL}/christmas-gifts/",
    f"{BASE_URL}/valentines-gifts/",
    f"{BASE_URL}/anniversary-gifts/",
    f"{BASE_URL}/white-elephant-gifts/",
    f"{BASE_URL}/office-gifts/",
    f"{BASE_URL}/kitchen-gadgets/",
    f"{BASE_URL}/outdoor-gifts/",
    f"{BASE_URL}/gaming-gifts/",
    f"{BASE_URL}/beer-gifts/",
    f"{BASE_URL}/wine-gifts/",
    f"{BASE_URL}/coffee-gifts/",
    f"{BASE_URL}/cat-gifts/",
    f"{BASE_URL}/dog-gifts/",
    f"{BASE_URL}/fitness-gifts/",
    f"{BASE_URL}/travel-gifts/",
    f"{BASE_URL}/music-gifts/",
    f"{BASE_URL}/art-gifts/",
    f"{BASE_URL}/book-gifts/",
    f"{BASE_URL}/science-gifts/",
    f"{BASE_URL}/nerdy-gifts/",
    f"{BASE_URL}/geek-gifts/",
    f"{BASE_URL}/star-wars-gifts/",
    f"{BASE_URL}/harry-potter-gifts/",
    f"{BASE_URL}/anime-gifts/",
    f"{BASE_URL}/minecraft-gifts/",
    f"{BASE_URL}/fortnite-gifts/",
]


def extract_products_from_html(html: str, source_url: str, category_hint: str = "") -> list[dict]:
    """Extrahiert Produkte aus dem geparsten HTML."""
    soup = BeautifulSoup(html, "lxml")
    products = []

    # Methode 1: Standard Produkt-Cards (JSON-LD / og-tags)
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string)
            if isinstance(data, list):
                items = data
            elif isinstance(data, dict):
                items = [data]
            else:
                continue
            for item in items:
                if item.get("@type") in ("Product", "ItemPage"):
                    p = {
                        "name": item.get("name", ""),
                        "description": item.get("description", ""),
                        "price": "",
                        "currency": "USD",
                        "category": category_hint,
                        "image_url": "",
                        "affiliate_url": source_url,
                        "amazon_url": "",
                        "source_page": source_url,
                        "brand": item.get("brand", {}).get("name", "") if isinstance(item.get("brand"), dict) else "",
                        "rating": "",
                        "in_stock": "Ja",
                    }
                    offers = item.get("offers", {})
                    if isinstance(offers, dict):
                        p["price"] = offers.get("price", "")
                        p["currency"] = offers.get("priceCurrency", "USD")
                    if item.get("image"):
                        img = item["image"]
                        p["image_url"] = img[0] if isinstance(img, list) else img
                    if item.get("aggregateRating"):
                        r = item["aggregateRating"]
                        p["rating"] = f"{r.get('ratingValue','')}/5 ({r.get('reviewCount','')} Bewertungen)"
                    if p["name"]:
                        products.append(p)
        except Exception:
            pass

    # Methode 2: Product-Cards via CSS-Klassen
    card_selectors = [
        "article.product",
        "div.product-card",
        "div.product-item",
        "div[class*='product']",
        "li.product",
        "div.item-card",
        "div.gift-card",
        "div[class*='card']",
        "div[data-product]",
    ]
    for selector in card_selectors:
        cards = soup.select(selector)
        for card in cards:
            name_el = card.select_one("h2, h3, h4, .product-title, .title, [class*='title'], [class*='name']")
            desc_el = card.select_one("p, .description, .desc, [class*='desc']")
            price_el = card.select_one(".price, [class*='price'], span[class*='cost']")
            link_el = card.select_one("a[href]")
            img_el = card.select_one("img")

            name = name_el.get_text(strip=True) if name_el else ""
            if not name or len(name) < 3:
                continue

            link = ""
            if link_el:
                href = link_el.get("href", "")
                if href.startswith("http"):
                    link = href
                elif href.startswith("/"):
                    link = BASE_URL + href

            amazon_url = ""
            if "amazon" in link.lower():
                amazon_url = link

            p = {
                "name": name,
                "description": desc_el.get_text(strip=True)[:300] if desc_el else "",
                "price": price_el.get_text(strip=True) if price_el else "",
                "currency": "USD",
                "category": category_hint,
                "image_url": img_el.get("src", img_el.get("data-src", "")) if img_el else "",
                "affiliate_url": link or source_url,
                "amazon_url": amazon_url,
                "source_page": source_url,
                "brand": "",
                "rating": "",
                "in_stock": "Ja",
            }
            products.append(p)
        if products:
            break

    # Methode 3: Alle Links die auf amazon.com zeigen
    if not products:
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "amazon" not in href.lower() and "thisiswhyimbroke" not in href.lower():
                continue

            # Produktname aus umliegendem Text
            name = a.get_text(strip=True)
            if not name:
                parent = a.parent
                if parent:
                    name = parent.get_text(strip=True)[:80]
            if not name or len(name) < 5:
                continue

            img = a.find("img")
            img_url = img.get("src", "") if img else ""

            p = {
                "name": name[:120],
                "description": "",
                "price": "",
                "currency": "USD",
                "category": category_hint,
                "image_url": img_url,
                "affiliate_url": href,
                "amazon_url": href if "amazon" in href.lower() else "",
                "source_page": source_url,
                "brand": "",
                "rating": "",
                "in_stock": "Ja",
            }
            products.append(p)

    # Methode 4: Open Graph Meta-Tags (Einzelprodukt-Seiten)
    if not products:
        og_title = soup.find("meta", property="og:title")
        og_desc = soup.find("meta", property="og:description")
        og_img = soup.find("meta", property="og:image")
        og_url = soup.find("meta", property="og:url")

        name = og_title["content"] if og_title and og_title.get("content") else ""
        if name and len(name) > 3:
            p = {
                "name": name,
                "description": og_desc["content"] if og_desc and og_desc.get("content") else "",
                "price": "",
                "currency": "USD",
                "category": category_hint,
                "image_url": og_img["content"] if og_img and og_img.get("content") else "",
                "affiliate_url": og_url["content"] if og_url and og_url.get("content") else source_url,
                "amazon_url": "",
                "source_page": source_url,
                "brand": "",
                "rating": "",
                "in_stock": "Ja",
            }
            # Preis aus Seitentext
            price_match = re.search(r'\$[\d,]+\.?\d*', soup.get_text())
            if price_match:
                p["price"] = price_match.group()
            # Amazon-Link suchen
            for a in soup.find_all("a", href=True):
                if "amazon" in a["href"].lower():
                    p["amazon_url"] = a["href"]
                    break
            products.append(p)

    return products


def extract_links_from_html(html: str, base: str) -> list[str]:
    """Findet alle internen Links zu Produkt- und Kategorieseiten."""
    soup = BeautifulSoup(html, "lxml")
    links = set()
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if href.startswith("/") and not href.startswith("//"):
            full = base + href
        elif href.startswith(base):
            full = href
        else:
            continue
        # Nur Produkt- und Kategorieseiten (keine accounts, login, etc.)
        skip = ["#", "login", "signup", "account", "cart", "checkout", "privacy", "dmca",
                "about", "contact", "mailto", "javascript", ".pdf", ".jpg", ".png"]
        if any(s in full.lower() for s in skip):
            continue
        links.add(full.rstrip("/") + "/")
    return list(links)


def guess_category(url: str, name: str) -> str:
    """Leitet Kategorie aus URL oder Produktname ab."""
    text = (url + " " + name).lower()
    mapping = {
        "Küche & Kochen": ["kitchen", "cook", "food", "chef", "bake", "coffee", "wine", "beer", "pizza", "pasta"],
        "Tech & Gadgets": ["tech", "gadget", "electronic", "smart", "usb", "wireless", "bluetooth", "led", "digital"],
        "Gaming": ["gaming", "game", "gamer", "controller", "xbox", "playstation", "nintendo", "arcade"],
        "Outdoor & Camping": ["outdoor", "camp", "survival", "hiking", "adventure", "fishing", "hunting"],
        "Deko & Wohnen": ["decor", "lamp", "light", "wall", "home", "room", "desk", "furniture"],
        "Wellness & Beauty": ["wellness", "beauty", "massage", "yoga", "meditation", "skin", "hair"],
        "Fashion & Accessoires": ["fashion", "clothing", "shirt", "hat", "bag", "wallet", "shoe", "sock"],
        "Tiere & Haustiere": ["pet", "dog", "cat", "animal"],
        "Wissenschaft & Bildung": ["science", "education", "book", "learn", "school", "stem"],
        "Sport & Fitness": ["sport", "fitness", "gym", "workout", "exercise", "bike", "run"],
        "Musik & Audio": ["music", "audio", "speaker", "headphone", "guitar", "piano"],
        "Partys & Spaß": ["party", "fun", "joke", "prank", "weird", "funny", "bizarre", "novelty", "gift"],
        "Kinder & Familie": ["kids", "children", "baby", "family", "teen"],
        "Reise": ["travel", "luggage", "passport", "airplane"],
        "Auto & Fahrzeug": ["car", "auto", "vehicle", "driving"],
        "Büro & Arbeit": ["office", "work", "desk", "organiz"],
        "Star Wars & Sci-Fi": ["star wars", "darth", "yoda", "jedi", "sci-fi", "space", "galaxy"],
        "Harry Potter": ["harry potter", "hogwarts", "wizard"],
        "Anime & Manga": ["anime", "manga", "japanese"],
    }
    for cat, keywords in mapping.items():
        if any(kw in text for kw in keywords):
            return cat
    return "Sonstiges / Lustige Geschenke"


async def crawl_page(crawler: AsyncWebCrawler, url: str, config: CrawlerRunConfig) -> tuple[str, str]:
    """Crawlt eine Seite und gibt (html, final_url) zurück."""
    try:
        result = await crawler.arun(url=url, config=config)
        if result and result.success:
            return result.html or "", result.url or url
        return "", url
    except Exception as e:
        print(f"  ⚠ Fehler bei {url}: {e}")
        return "", url


async def main():
    print("=" * 60)
    print("thisiswhyimbroke.com Crawler")
    print(f"Gestartet: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    browser_cfg = BrowserConfig(
        headless=True,
        verbose=False,
        extra_args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage"],
    )

    run_cfg = CrawlerRunConfig(
        wait_for="css:body",
        page_timeout=20000,
        delay_before_return_html=3.0,
        js_code="""
        // Warte auf dynamischen Content
        await new Promise(r => setTimeout(r, 2500));
        // Scrolle um lazy-load zu triggern
        window.scrollTo(0, document.body.scrollHeight / 2);
        await new Promise(r => setTimeout(r, 800));
        window.scrollTo(0, document.body.scrollHeight);
        await new Promise(r => setTimeout(r, 800));
        """,
        cache_mode="bypass",
    )

    all_products = []
    seen_names = set()
    seen_urls = set()
    visited_pages = set()
    pages_to_visit = list(SEED_URLS)

    async with AsyncWebCrawler(config=browser_cfg) as crawler:

        # Phase 1: Kategorieseiten crawlen & Produkt-URLs sammeln
        print("\n[Phase 1] Kategorieseiten crawlen...")
        product_urls = set()
        category_url_map = {}

        batch_size = 5
        for i in range(0, min(len(pages_to_visit), 50), batch_size):
            batch = pages_to_visit[i:i+batch_size]
            tasks = [crawl_page(crawler, url, run_cfg) for url in batch]
            results = await asyncio.gather(*tasks)

            for url, (html, final_url) in zip(batch, results):
                visited_pages.add(url)
                if not html:
                    print(f"  ✗ Kein Inhalt: {url}")
                    continue

                cat = url.replace(BASE_URL, "").strip("/").replace("-", " ").title()
                prods = extract_products_from_html(html, final_url, cat)
                links = extract_links_from_html(html, BASE_URL)

                # Produkte direkt sammeln
                added = 0
                for p in prods:
                    name_key = p["name"].lower().strip()
                    if name_key and name_key not in seen_names and len(name_key) > 4:
                        seen_names.add(name_key)
                        p["category"] = guess_category(url, p["name"])
                        all_products.append(p)
                        added += 1

                # Neue Links einsammeln (Produkt-Seiten identifizieren)
                for link in links:
                    clean = link.rstrip("/")
                    if clean not in visited_pages and clean not in seen_urls:
                        # Kategorieseiten haben "-gifts", "-stuff", "-gadgets" im Namen
                        # Produktseiten haben ein einzelnes Wort/kurze Phrase
                        path = link.replace(BASE_URL, "").strip("/")
                        segments = path.split("/")
                        if len(segments) == 1 and len(path) > 3:
                            # Könnte Produktseite oder Kategorieseite sein
                            if any(kw in path for kw in ["gifts", "stuff", "gadgets", "items", "things", "products", "guide"]):
                                if link not in pages_to_visit:
                                    pages_to_visit.append(link)
                                    category_url_map[link] = cat
                            else:
                                product_urls.add(link)

                print(f"  ✓ {url.replace(BASE_URL,'')} → {added} Produkte, {len(links)} Links gefunden")

            print(f"  Zwischenstand: {len(all_products)} Produkte gesammelt")
            await asyncio.sleep(1)

        # Phase 2: Weitere Kategorieseiten die wir gefunden haben
        extra_cats = [u for u in pages_to_visit if u not in visited_pages][:100]
        if extra_cats:
            print(f"\n[Phase 2] {len(extra_cats)} weitere Kategorieseiten...")
            for i in range(0, len(extra_cats), batch_size):
                batch = extra_cats[i:i+batch_size]
                tasks = [crawl_page(crawler, url, run_cfg) for url in batch]
                results = await asyncio.gather(*tasks)

                for url, (html, final_url) in zip(batch, results):
                    visited_pages.add(url)
                    if not html:
                        continue
                    cat = url.replace(BASE_URL, "").strip("/").replace("-", " ").title()
                    prods = extract_products_from_html(html, final_url, cat)
                    for p in prods:
                        name_key = p["name"].lower().strip()
                        if name_key and name_key not in seen_names and len(name_key) > 4:
                            seen_names.add(name_key)
                            p["category"] = guess_category(url, p["name"])
                            all_products.append(p)
                    # Neue Produkt-Links sammeln
                    links = extract_links_from_html(html, BASE_URL)
                    for link in links:
                        path = link.replace(BASE_URL, "").strip("/")
                        if path and "/" not in path and link not in visited_pages:
                            product_urls.add(link)

                print(f"  Fortschritt: {len(all_products)} Produkte, {len(product_urls)} Produkt-URLs")
                await asyncio.sleep(0.5)

        # Phase 3: Einzelne Produktseiten crawlen (bis wir 1000 haben)
        remaining = max(0, 1000 - len(all_products))
        product_url_list = list(product_urls - visited_pages)[:remaining + 200]

        if product_url_list:
            print(f"\n[Phase 3] {len(product_url_list)} Produktseiten crawlen...")
            prod_cfg = CrawlerRunConfig(
                wait_for="css:body",
                page_timeout=15000,
                delay_before_return_html=2.0,
                cache_mode="bypass",
            )
            for i in range(0, len(product_url_list), batch_size):
                if len(all_products) >= 1000:
                    break
                batch = product_url_list[i:i+batch_size]
                tasks = [crawl_page(crawler, url, prod_cfg) for url in batch]
                results = await asyncio.gather(*tasks)

                for url, (html, final_url) in zip(batch, results):
                    if not html:
                        continue
                    path = url.replace(BASE_URL, "").strip("/")
                    cat_hint = path.replace("-", " ").title()
                    prods = extract_products_from_html(html, final_url, cat_hint)
                    for p in prods[:1]:  # Pro Produktseite max 1
                        name_key = p["name"].lower().strip()
                        if name_key and name_key not in seen_names and len(name_key) > 4:
                            seen_names.add(name_key)
                            p["category"] = guess_category(url, p["name"])
                            all_products.append(p)

                if i % 25 == 0:
                    print(f"  {i}/{len(product_url_list)} gecrawlt → {len(all_products)} Produkte")
                await asyncio.sleep(0.3)

    print(f"\n✓ Gesamt gesammelt: {len(all_products)} Produkte")

    # Deduplizierung nochmals
    seen = set()
    clean_products = []
    for p in all_products:
        key = p["name"].lower().strip()[:60]
        if key not in seen and len(key) > 4:
            seen.add(key)
            clean_products.append(p)
    all_products = clean_products
    print(f"✓ Nach Deduplizierung: {len(all_products)} Produkte")

    # Sortieren nach Kategorie
    all_products.sort(key=lambda x: (x.get("category", ""), x.get("name", "")))

    # Excel erstellen
    create_excel(all_products)


def create_excel(products: list[dict]):
    print("\n[Excel] Erstelle Excel-Datei...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produkte"

    headers = [
        "Nr.", "Produktname", "Kategorie", "Beschreibung",
        "Preis (USD)", "Amazon DE Suchbegriff",
        "Affiliate URL", "Amazon URL", "Bild URL",
        "Marke", "Bewertung", "Auf Lager", "Quellseite"
    ]

    header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[1].height = 28

    cat_colors = {
        "Küche & Kochen": "FFF3E0",
        "Tech & Gadgets": "E3F2FD",
        "Gaming": "EDE7F6",
        "Outdoor & Camping": "E8F5E9",
        "Deko & Wohnen": "F1F8E9",
        "Wellness & Beauty": "FCE4EC",
        "Fashion & Accessoires": "FFFDE7",
        "Tiere & Haustiere": "F3E5F5",
        "Wissenschaft & Bildung": "E0F7FA",
        "Sport & Fitness": "E8EAF6",
        "Musik & Audio": "FBE9E7",
        "Partys & Spaß": "FFF8E1",
        "Kinder & Familie": "E0F2F1",
        "Reise": "E8EAF6",
        "Auto & Fahrzeug": "ECEFF1",
        "Büro & Arbeit": "F9FBE7",
        "Star Wars & Sci-Fi": "E8EAF6",
        "Harry Potter": "EDE7F6",
        "Anime & Manga": "FCE4EC",
        "Sonstiges / Lustige Geschenke": "FAFAFA",
    }

    for row_i, p in enumerate(products, 2):
        cat = p.get("category", "Sonstiges / Lustige Geschenke")
        fill_color = cat_colors.get(cat, "FAFAFA")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # Amazon DE Suchbegriff ableiten
        name = p.get("name", "")
        amazon_de_search = re.sub(r'[^\w\s]', ' ', name).strip()[:60]

        values = [
            row_i - 1,
            name,
            cat,
            p.get("description", "")[:300],
            p.get("price", ""),
            amazon_de_search,
            p.get("affiliate_url", ""),
            p.get("amazon_url", ""),
            p.get("image_url", ""),
            p.get("brand", ""),
            p.get("rating", ""),
            p.get("in_stock", "Ja"),
            p.get("source_page", ""),
        ]

        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_i, column=col, value=val)
            c.fill = fill
            c.border = border
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 1:
                c.font = Font(bold=True)

    col_widths = [5, 35, 25, 50, 12, 35, 45, 45, 40, 18, 20, 10, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Statistik-Sheet
    ws2 = wb.create_sheet("Kategorien")
    from collections import Counter
    cat_count = Counter(p.get("category", "Unbekannt") for p in products)
    ws2.cell(row=1, column=1, value="Kategorie").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Anzahl Produkte").font = Font(bold=True)
    for i, (cat, count) in enumerate(sorted(cat_count.items(), key=lambda x: -x[1]), 2):
        ws2.cell(row=i, column=1, value=cat)
        ws2.cell(row=i, column=2, value=count)
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 18

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    outpath = OUTPUT_DIR / f"thisiswhyimbroke_produkte_{ts}.xlsx"
    wb.save(outpath)
    print(f"✓ Excel gespeichert: {outpath}")
    print(f"✓ {len(products)} Produkte in {len(set(p.get('category') for p in products))} Kategorien")
    return outpath


if __name__ == "__main__":
    asyncio.run(main())
